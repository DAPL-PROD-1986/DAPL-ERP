

# import frappe

# @frappe.whitelist()
# def get_bom_items_custom(bom, warehouse=None, qty=1, fetch_exploded=1, company=None):

#     from erpnext.manufacturing.doctype.bom.bom import get_bom_items as core_method

#     items = core_method(
#         bom=bom,
#         qty=qty,
#         fetch_exploded=fetch_exploded,
#         company=company
#     )

#     if not items:
#         return items

#     # ✅ FETCH ALL BOM ITEMS (ALL LEVELS)
#     all_bom_items = frappe.get_all(
#         "BOM Item",
#         fields=[
#             "item_code",
#             "custom_item_group",
#             "custom_length",
#             "custom_width",
#             "custom_thickness",
#             "custom_density",
#             "custom_outer_diameter",
#             "custom_inner_diameter",
#             "custom_wall_thickness",
#             "custom_kilogramskgs",
#             "custom_total_weight"
#         ]
#     )

#     # 🔥 CREATE GLOBAL MAP (ITEM_CODE BASED)
#     item_map = {}

#     for d in all_bom_items:
#         # keep latest / non-zero values
#         if d.item_code not in item_map:
#             item_map[d.item_code] = d
#         else:
#             # prefer row with values
#             if any([
#                 d.custom_length,
#                 d.custom_width,
#                 d.custom_thickness,
#                 d.custom_density
#             ]):
#                 item_map[d.item_code] = d

#     # 🔁 MAP TO RESULT ITEMS
#     for item in items:

#         bom_item = item_map.get(item.get("item_code"))

#         if not bom_item:
#             continue

#         # ✅ assign values
#         item["item_group"] = bom_item.custom_item_group or item.get("item_group")

#         item["custom_length"] = bom_item.custom_length
#         item["custom_width"] = bom_item.custom_width
#         item["custom_thickness"] = bom_item.custom_thickness
#         item["custom_density"] = bom_item.custom_density

#         item["custom_outer_diameter"] = bom_item.custom_outer_diameter
#         item["custom_inner_diameter"] = bom_item.custom_inner_diameter
#         item["custom_wall_thickness"] = bom_item.custom_wall_thickness

#         item["custom_kilogramskgs"] = bom_item.custom_kilogramskgs
#         item["custom_total_weight"] = bom_item.custom_total_weight

#     return items




# import frappe

# def get_bom_map(bom):
#     """Recursively collect BOM Items from all levels"""

#     bom_map = {}

#     def fetch(bom_no):
#         items = frappe.get_all(
#             "BOM Item",
#             filters={"parent": bom_no},
#             fields=[
#                 "item_code",
#                 "bom_no",
#                 "custom_item_group",
#                 "custom_length",
#                 "custom_width",
#                 "custom_thickness",
#                 "custom_density",
#                 "custom_outer_diameter",
#                 "custom_inner_diameter",
#                 "custom_wall_thickness",
#                 "custom_kilogramskgs",
#                 "custom_total_weight"
#             ]
#         )

#         for d in items:
#             bom_map.setdefault(d.item_code, []).append(d)

#             # 🔁 GO DEEP (SUB BOM)
#             if d.bom_no:
#                 fetch(d.bom_no)

#     fetch(bom)
#     return bom_map


# @frappe.whitelist()
# def get_bom_items_custom(bom, warehouse=None, qty=1, fetch_exploded=1, company=None):

#     from erpnext.manufacturing.doctype.bom.bom import get_bom_items as core_method

#     items = core_method(
#         bom=bom,
#         qty=qty,
#         fetch_exploded=fetch_exploded,
#         company=company
#     )

#     if not items:
#         return items

#     # 🔥 GET FULL BOM STRUCTURE (ALL LEVELS)
#     bom_map = get_bom_map(bom)

#     # 🔁 MAP VALUES
#     for item in items:

#         rows = bom_map.get(item.get("item_code"))

#         if not rows:
#             continue

#         # 👉 pick row with actual values
#         matched = None

#         for r in rows:
#             if any([
#                 r.custom_length,
#                 r.custom_width,
#                 r.custom_thickness,
#                 r.custom_density
#             ]):
#                 matched = r
#                 break

#         if not matched:
#             matched = rows[0]

#         # ✅ MAP
#         item["item_group"] = matched.custom_item_group or item.get("item_group")

#         item["custom_length"] = matched.custom_length
#         item["custom_width"] = matched.custom_width
#         item["custom_thickness"] = matched.custom_thickness
#         item["custom_density"] = matched.custom_density

#         item["custom_outer_diameter"] = matched.custom_outer_diameter
#         item["custom_inner_diameter"] = matched.custom_inner_diameter
#         item["custom_wall_thickness"] = matched.custom_wall_thickness

#         item["custom_kilogramskgs"] = matched.custom_kilogramskgs
#         item["custom_total_weight"] = matched.custom_total_weight

#     return items




# import frappe
# from frappe import _
# from frappe.utils import flt

# def get_bom_map(bom):
#     """Recursively collect BOM Items from all levels"""

#     bom_map = {}

#     def fetch(bom_no):
#         items = frappe.get_all(
#             "BOM Item",
#             filters={"parent": bom_no},
#             fields=[
#                 "item_code",
#                 "bom_no",
#                 "custom_item_group",
#                 "custom_length",
#                 "custom_width",
#                 "custom_thickness",
#                 "custom_density",
#                 "custom_outer_diameter",
#                 "custom_inner_diameter",
#                 "custom_wall_thickness",
#                 "custom_kilogramskgs",
#                 "custom_total_weight"
#             ]
#         )

#         for d in items:
#             bom_map.setdefault(d.item_code, []).append(d)

#             # 🔁 RECURSION (SUB BOM)
#             if d.bom_no:
#                 fetch(d.bom_no)

#     fetch(bom)
#     return bom_map


# @frappe.whitelist()
# def get_bom_items_custom(bom, warehouse=None, qty=1, fetch_exploded=1, company=None):

#     from erpnext.manufacturing.doctype.bom.bom import get_bom_items as core_method

#     # ✅ STEP 1: Get exploded items (final RM list)
#     items = core_method(
#         bom=bom,
#         qty=qty,
#         fetch_exploded=fetch_exploded,
#         company=company
#     )

#     if not items:
#         return items

#     # ✅ STEP 2: Get FULL BOM STRUCTURE (ALL LEVELS)
#     bom_map = get_bom_map(bom)

#     # ✅ STEP 3: Get Item Group Parent Mapping
#     item_groups = frappe.get_all(
#         "Item Group",
#         fields=["name", "parent_item_group"]
#     )

#     group_map = {g.name: g.parent_item_group for g in item_groups}

#     final_items = []

#     # 🔁 STEP 4: MAP + FILTER
#     for item in items:

#         item_code = item.get("item_code")
#         item_group = item.get("item_group")

#         if not item_group:
#             continue

#         parent_group = group_map.get(item_group)

#         # ✅ FILTER ONLY PURCHASE PLAN
#         if parent_group != "Purchase Plan":
#             continue

#         rows = bom_map.get(item_code)

#         if not rows:
#             continue

#         # 🔥 PICK BEST MATCH (IMPORTANT FIX)
#         matched = None

#         for r in rows:
#             if any([
#                 r.custom_length,
#                 r.custom_width,
#                 r.custom_thickness,
#                 r.custom_density,
#                 r.custom_outer_diameter
#             ]):
#                 matched = r
#                 break

#         if not matched:
#             matched = rows[0]

#         # ✅ FINAL MAPPING
#         item["item_group"] = matched.custom_item_group or item.get("item_group")

#         item["custom_length"] = matched.custom_length or 0
#         item["custom_width"] = matched.custom_width or 0
#         item["custom_thickness"] = matched.custom_thickness or 0
#         item["custom_density"] = matched.custom_density or 0

#         item["custom_outer_diameter"] = matched.custom_outer_diameter or 0
#         item["custom_inner_diameter"] = matched.custom_inner_diameter or 0
#         item["custom_wall_thickness"] = matched.custom_wall_thickness or 0

#         item["custom_kilogramskgs"] = matched.custom_kilogramskgs or 0
#         item["custom_total_weight"] = matched.custom_total_weight or 0

#         final_items.append(item)

#     return final_items




import frappe
import math
import io
import json

from frappe import _
from frappe.utils import flt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ========================== BOM MAP =================================

def get_bom_map(bom):
    """Recursively collect BOM Items from all levels."""

    bom_map = {}

    def fetch(bom_no):
        items = frappe.get_all(
            "BOM Item",
            filters={"parent": bom_no},
            fields=[
                "item_code",
                "bom_no",
                "custom_item_group",
                "custom_length",
                "custom_width",
                "custom_thickness",
                "custom_density",
                "custom_outer_diameter",
                "custom_inner_diameter",
                "custom_wall_thickness",
                "custom_kilogramskgs",
                "custom_total_weight"
            ]
        )

        for d in items:
            bom_map.setdefault(d.item_code, []).append(d)

            if d.bom_no:
                fetch(d.bom_no)

    fetch(bom)
    return bom_map


# ==================== GET ITEMS FROM BOM =====================================
@frappe.whitelist()
def get_bom_items_custom(
    bom,
    warehouse=None,
    qty=1,
    fetch_exploded=1,
    company=None
):

    from erpnext.manufacturing.doctype.bom.bom import get_bom_items as core_method
    items = core_method(
        bom=bom,
        qty=qty,
        fetch_exploded=fetch_exploded,
        company=company
    )

    if not items:
        return items

    bom_map = get_bom_map(bom)
    item_groups = frappe.get_all("Item Group", fields=["name", "parent_item_group"])

    group_map = {
        g.name: g.parent_item_group
        for g in item_groups
    }

    final_items = []
    for item in items:
        item_code = item.get("item_code")
        item_group = item.get("item_group")

        if not item_group:
            continue

        parent_group = group_map.get(item_group)
        if parent_group != "Purchase Plan":
            continue

        rows = bom_map.get(item_code)
        if not rows:
            continue

        matched = None
        for r in rows:
            if any([
                r.custom_length,
                r.custom_width,
                r.custom_thickness,
                r.custom_density,
                r.custom_outer_diameter
            ]):
                matched = r
                break

        if not matched:
            matched = rows[0]

        item["item_group"] = (matched.custom_item_group or item.get("item_group"))
        item["custom_length"] = (matched.custom_length or 0)
        item["custom_width"] = (matched.custom_width or 0)
        item["custom_thickness"] = (matched.custom_thickness or 0)
        item["custom_density"] = (matched.custom_density or 0)
        item["custom_outer_diameter"] = (matched.custom_outer_diameter or 0)
        item["custom_inner_diameter"] = (matched.custom_inner_diameter or 0)
        item["custom_kilogramskgs"] = (matched.custom_kilogramskgs or 0)
        item["custom_total_weight"] = (matched.custom_total_weight or 0)

        final_items.append(item)
    return final_items


# ======================== CALCULATE MATERIAL REQUEST ITEM ====================================
@frappe.whitelist()
def recalc_item(item):

    if isinstance(item, str):
        item = json.loads(item)

    return calculate_values(item)


def calculate_values(item):

    def get(key):
        if isinstance(item, dict):
            return item.get(key)

        return getattr(item, key, None)

    def setv(key, value):
        if isinstance(item, dict):
            item[key] = value
        else:
            setattr(item, key, value)

    density = flt(get("custom_density"))
    qty = flt(get("qty"))

    item_group = get("item_group")
    shape = get("custom_shape")

    length = flt(get("custom_length"))
    width = flt(get("custom_width"))
    thickness = flt(get("custom_thickness"))
    outer_diameter = flt(get("custom_outer_diameter"))
    inner_diameter = flt(get("custom_inner_diameter"))

    pi = math.pi
    base_weight = 0

    # ======================= MANUAL ENTRY MODE ============================
    if shape == "N/A":
        manual_kg = flt(get("custom_kilogramskgs"))
        manual_total_weight = flt(get("custom_total_weight"))

        # Kgs Per Unit -> Total Weight
        if manual_kg > 0 and qty > 0:
            manual_total_weight = (manual_kg * qty)

        # Total Weight -> Kgs Per Unit
        elif manual_total_weight > 0 and qty > 0:
            manual_kg = (manual_total_weight / qty)

        setv("custom_kilogramskgs", flt(manual_kg, 4))
        setv("custom_total_weight", flt(manual_total_weight, 4))

        # ======================= MANUAL METER ENTRY ==================================
        manual_mtr_per_unit = flt(get("custom_mtr_per_unit"))
        manual_total_mtr = flt(get("custom_total_mtr"))

        # Mtr Per Unit -> Total Mtr
        if manual_mtr_per_unit > 0 and qty > 0:
            manual_total_mtr = (manual_mtr_per_unit * qty)

        # Total Mtr -> Mtr Per Unit
        elif manual_total_mtr > 0 and qty > 0:
            manual_mtr_per_unit = (manual_total_mtr / qty)

        setv("custom_mtr_per_unit", flt(manual_mtr_per_unit, 4))
        setv("custom_total_mtr", flt(manual_total_mtr, 4))
        return item

    # ========================= AUTO CALCULATION MODE ===================================
    if not density:
        setv("custom_kilogramskgs", 0)
        setv("custom_total_weight", 0)

    else:
        # ===================== PLATES ================================
        if item_group == "Plates":

            # Rectangle
            if (shape == "Rectangle" and length and width and thickness):
                base_weight = (length * width * thickness * density) / 1000000

            # Circle
            elif (shape == "Circle" and outer_diameter and thickness):
                base_weight = (pi * (outer_diameter / 2) ** 2 * thickness * density) / 1000000

            # Hollow
            elif shape == "Hollow":
                calculated_od = (outer_diameter if outer_diameter else inner_diameter + (2 * thickness))

                if (calculated_od and inner_diameter and length):
                    base_weight = (pi * ((calculated_od / 2) ** 2 - (inner_diameter / 2) ** 2) * length * density) / 1000000

        # ======================== PIPES / TUBES ================================
        elif item_group in ["Pipes", "Tubes"]:
            if (shape == "Hollow" and outer_diameter and thickness and length):
                calculated_id = (outer_diameter - (2 * thickness))
                if calculated_id > 0:

                    base_weight = (pi * ((outer_diameter / 2) ** 2 - (calculated_id / 2) ** 2) * length * density) / 1000000

        # ========================= FORGINGS ==================================
        elif item_group == "Forgings":

            # Hollow
            if (shape == "Hollow" and outer_diameter and thickness and length):
                calculated_id = (outer_diameter - (2 * thickness))

                if calculated_id > 0:
                    base_weight = (pi * ((outer_diameter / 2) ** 2 - (calculated_id / 2) ** 2) * length * density) / 1000000

            # Circle
            elif (shape == "Circle" and outer_diameter and thickness):
                base_weight = (pi * (outer_diameter / 2) ** 2 * thickness * density) / 1000000

        # ===================== RODS ==============================
        elif item_group == "Rods":
            if (shape == "Circle" and outer_diameter and length):
                base_weight = (pi * (outer_diameter / 2) ** 2 * length * density) / 1000000

        # ==================== FLANGES / RINGS ===============================
        elif item_group in ["Flanges", "Rings"]:
            if (outer_diameter and inner_diameter and thickness):
                base_weight = (pi * ((outer_diameter / 2) ** 2 - (inner_diameter / 2) ** 2) * thickness * density) / 1000000

        # ====================== FINAL WEIGHT ==================================
        kg_per_unit = flt(base_weight, 4)
        total_weight = flt(qty * kg_per_unit, 4)

        setv("custom_kilogramskgs", kg_per_unit)
        setv("custom_total_weight", total_weight)

    # ========================= METER CALCULATION =================================
    manual_mtr_per_unit = flt(get("custom_mtr_per_unit"))
    manual_total_mtr = flt(get("custom_total_mtr"))

    # ========================= MANUAL ENTRY MODE =========================
    if shape == "N/A":

        # Mtr Per Unit → Total Mtr
        if manual_mtr_per_unit > 0 and qty > 0:
            manual_total_mtr = manual_mtr_per_unit * qty

        # Total Mtr → Mtr Per Unit
        elif manual_total_mtr > 0 and qty > 0:
            manual_mtr_per_unit = manual_total_mtr / qty

        setv("custom_mtr_per_unit", flt(manual_mtr_per_unit, 4))
        setv("custom_total_mtr", flt(manual_total_mtr, 4))

    # ========================= AUTO ENTRY MODE =========================
    else:
        mtr_per_unit = flt(length / 1000, 4)
        total_mtr = flt(mtr_per_unit * qty, 4)

        setv("custom_mtr_per_unit", mtr_per_unit)
        setv("custom_total_mtr",total_mtr)

    return item

# ========================== UPLOAD MATERIAL REQUEST EXCEL ====================================

@frappe.whitelist()
def upload_material_request_excel(file_url=None):

    if not file_url:
        file_url = frappe.form_dict.get("file_url")

    if not file_url:
        frappe.throw(_("Excel file was not uploaded."))

    file_doc = frappe.get_doc("File", {
            "file_url": file_url
        })

    file_path = file_doc.get_full_path()

    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return []

    # ========================= PROFESSIONAL EXCEL HEADERS ===================================
    expected_headers = [
        "Item Code", "Item Name", "Item Group", "Shape", "Material Type", "Qty", "UOM",
        "Length (mm)", "Width (mm)", "Thickness (mm)", "Outer Diameter (mm)", "Inner Diameter (mm)", "Density (kg/m³)",
        "Kgs Per Unit", "Total Weight", "Mtr Per Unit", "Total Mtr", "Description", "BOM No", "BOM Part No"
    ]

    headers = [str(value).strip()
        if value is not None
        else ""
        for value in rows[0]
    ]

    missing_columns = [column
        for column in expected_headers
        if column not in headers
    ]

    if missing_columns:
        frappe.throw(_("Missing columns: {0}").format(", ".join(missing_columns)))

    # ======================= FIELD MAP ===============================
    field_map = {
        "Item Code": "item_code",
        "Item Name": "item_name",
        "Item Group": "item_group",
        "Shape": "custom_shape",
        "Material Type": "custom_material_type",
        "Qty": "qty",
        "UOM": "uom",
        "Length (mm)": "custom_length",
        "Width (mm)": "custom_width",
        "Thickness (mm)": "custom_thickness",
        "Outer Diameter (mm)": "custom_outer_diameter",
        "Inner Diameter (mm)": "custom_inner_diameter",
        "Density (kg/m³)": "custom_density",
        "Kgs Per Unit": "custom_kilogramskgs",
        "Total Weight": "custom_total_weight",
        "Mtr Per Unit": "custom_mtr_per_unit",
        "Total Mtr": "custom_total_mtr",

        "Description": "description",
        "BOM No": "bom_no",
        "BOM Part No": "custom_bom_part_no"
    }

    column_index = {
        column: headers.index(column)
        for column in expected_headers
    }

    result = []

    # ======================= READ EXCEL DATA =================================
    for excel_row in rows[1:]:
        if not any(
            value is not None
            and str(value).strip()
            for value in excel_row
        ):
            continue

        def get_value(column):
            index = column_index[column]

            if index >= len(excel_row):
                return ""

            return excel_row[index]
        item_code = get_value("Item Code")

        if not item_code:
            continue

        item_code = str(item_code).strip()

        # =========================== VALIDATE ITEM ================================
        if not frappe.db.exists("Item", item_code):
            frappe.throw(_("Item {0} does not exist.").format(item_code))
        child = {}

        # ======================= MAP EXCEL -> FRAPPE =============================
        for excel_field, frappe_field in field_map.items():
            value = get_value(excel_field)
            if frappe_field in ["qty"]:
                value = flt(value)

            elif frappe_field in [
                "custom_length", "custom_width", "custom_thickness", "custom_outer_diameter", "custom_inner_diameter", "custom_density",
                "custom_kilogramskgs", "custom_total_weight", "custom_mtr_per_unit", "custom_total_mtr"
            ]:
                value = flt(value)

            elif value is None:
                value = ""

            else:
                value = str(value).strip()
            child[frappe_field] = value

            # Always set Conversion Factor = 1
        child["conversion_factor"] = 1

        # ====================== AUTO FETCH ITEM DATA ============================
        item_data = frappe.db.get_value("Item", item_code, [
                "item_name",
                "stock_uom",
                "item_group",
                "description",
                "default_bom",
                "custom_material_type",
                "custom_density",
                "custom_thickness"
            ], as_dict=True)

        if item_data:
            if not child.get("item_name"):
                child["item_name"] = (item_data.item_name or "")

            if not child.get("uom"):
                child["uom"] = (item_data.stock_uom or "")

            if not child.get("item_group"):
                child["item_group"] = (item_data.item_group or "")

            if not child.get("description"):
                child["description"] = item_data.description or ""

            if not child.get("bom_no"):
                child["bom_no"] = item_data.default_bom or ""

            if not child.get("custom_material_type"):
                child["custom_material_type"] = (item_data.custom_material_type or "")

            if not child.get("custom_density"):
                child["custom_density"] = (item_data.custom_density or 0)

            if not child.get("custom_thickness"):
                child["custom_thickness"] = (item_data.custom_thickness or 0)

        # ======================= CALCULATE VALUES ===============================
        child = calculate_values(child)
        result.append(child)
    return result


# =========================== DOWNLOAD MATERIAL REQUEST TEMPLATE ======================================
@frappe.whitelist()
def download_material_request_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Material Request"

    # ======================== HEADERS ===============================
    headers = [
        "Item Code", "Item Name", "Item Group", "Shape", "Material Type", "Qty", "UOM",
        "Length (mm)", "Width (mm)", "Thickness (mm)", "Outer Diameter (mm)", "Inner Diameter (mm)", "Density (kg/m³)",
        "Kgs Per Unit", "Total Weight", "Mtr Per Unit", "Total Mtr", "Description", "BOM No", "BOM Part No"
    ]
    worksheet.append(headers)

    # ======================= HEADER STYLE ===========================
    header_fill = PatternFill(fill_type="solid", fgColor="DCFCE7")
    header_font = Font(bold=True, color="166534")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="B7D7C0")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    worksheet.row_dimensions[1].height = 35

    # ========================== COLUMN WIDTH ===================================
    column_widths = {
        "A": 22,
        "B": 25,
        "C": 20,
        "D": 15,
        "E": 20,
        "F": 12,
        "G": 12,
        "H": 17,
        "I": 17,
        "J": 19,
        "K": 23,
        "L": 23,
        "M": 21,
        "N": 17,
        "O": 18,
        "P": 17,
        "Q": 15,
        "R": 35,
        "S": 25,
        "T": 25
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # ====================== FREEZE HEADER ================================
    worksheet.freeze_panes = "A2"

    # ======================= FILTER ==================================
    worksheet.auto_filter.ref = (f"A1:T1")

    # ======================== CREATE FILE ================================
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": "Material_Request_Template.xlsx",
        "is_private": 0,
        "content": output.read()
    })

    file_doc.insert(ignore_permissions=True)
    return file_doc.file_url