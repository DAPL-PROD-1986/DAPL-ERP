# Copyright (c) 2026, maze and contributors
# For license information, please see license.txt

# import frappe
# from frappe.model.document import Document


# class StockItem(Document):
# 	pass


import frappe
import io
import json
import openpyxl
from frappe.model.document import Document
from math import pi
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from frappe.utils.file_manager import get_file
from frappe.utils.xlsxutils import make_xlsx


class StockItem(Document):

    def validate(self):
        self.calculate_stock_weights()

    def calculate_stock_weights(self):
        self.overall_avail_wgt_plate = 0
        self.overall_avail_wgt_amt_plate = 0

        self.overall_avail_wgt_tube = 0
        self.overall_avail_wgt_amt_tube = 0

        self.overall_avail_wgt_pipe = 0
        self.overall_avail_wgt_amt_pipe = 0
        self.overall_avail_mtr_pipe = 0
        self.overall_avail_mtr_tube = 0
        self.overall_avail_mtr_rod = 0

        self.overall_avail_wgt_rod = 0
        self.overall_avail_wgt_amt_rod = 0

        # Plates
        for row in self.plates or []:
            self.calculate_plate(row)
            self.overall_avail_wgt_plate += row.available_weight or 0
            self.overall_avail_wgt_amt_plate += row.available_weight_amount or 0

        # Tubes
        for row in self.tubes or []:
            self.calculate_tube(row)
            self.overall_avail_wgt_tube += row.available_weight or 0
            self.overall_avail_wgt_amt_tube += row.available_weight_amount or 0
            self.overall_avail_mtr_tube += row.available_meter or 0


        # Pipes
        for row in self.pipes or []:
            self.calculate_pipe(row)
            self.overall_avail_wgt_pipe += row.available_weight or 0
            self.overall_avail_wgt_amt_pipe += row.available_weight_amount or 0
            self.overall_avail_mtr_pipe += row.available_meter or 0

        # Rods
        for row in self.rods or []:
            self.calculate_rod(row)
            self.overall_avail_wgt_rod += row.available_weight or 0
            self.overall_avail_wgt_amt_rod += row.available_weight_amount or 0
            self.overall_avail_mtr_rod += row.available_meter or 0

    
    def calculate_plate(self, row):
        length = row.length or 0
        width = row.width or 0
        thickness = row.thickness or 0
        density = row.density or 0
        quantity = row.quantity or 0
        used_weight = row.used_weight or 0
        rate_per_kg = row.rate_per_kg or 0

        row.weight_per_item = (length * width * thickness * density) / 1000000
        row.actual_weight = (quantity * row.weight_per_item)
        row.available_weight = (row.actual_weight - used_weight)
        row.used_weight_percentage = ((row.actual_weight - row.available_weight) / row.actual_weight * 100) if row.actual_weight else 0
        row.actual_weight_amount = (row.actual_weight * rate_per_kg)
        row.available_weight_amount = (row.available_weight * rate_per_kg)


    def calculate_tube(self, row):
        length = row.length or 0
        outer_diameter = row.outer_diameter or 0
        thickness = row.thickness or 0
        density = row.density or 0
        quantity = row.quantity or 0
        used_quantity = row.used_quantity or 0
        used_weight = row.used_weight or 0
        used_meter = row.used_meter or 0
        rate_per_mtr = row.rate_per_mtr or 0

        if row.length and row.outer_diameter and row.thickness and row.density:
            od = row.outer_diameter
            inner_diameter = od - (2 * row.thickness)

            if inner_diameter > 0:
                row.weight_per_item = (pi * (((od / 2) ** 2) - ((inner_diameter / 2) ** 2)) * row.length * row.density) / 1000000
            else:
                row.weight_per_item = 0
        else:
            row.weight_per_item = 0
        
        row.balanced_quantity = ((row.quantity or 0) - (row.used_quantity or 0))
        row.actual_weight = (row.quantity or 1) * (row.weight_per_item or 0)
        row.available_weight = (row.actual_weight or 0) - (row.used_weight or 0)

        meter_per_item = length / 1000
        row.actual_meter = (quantity * meter_per_item)
        row.available_meter = max(row.actual_meter - used_meter,0)

        row.used_weight_percentage = (((row.actual_weight or 0) - (row.available_weight or 0)) / row.actual_weight * 100) if row.actual_weight else 0
        row.actual_weight_amount = (row.actual_weight or 0) * (row.rate_per_mtr or 0)
        row.available_weight_amount = (row.available_weight or 0) * (row.rate_per_mtr or 0)


    def calculate_pipe(self, row):
        length = row.length or 0
        outer_diameter = row.outer_diameter or 0
        thickness = row.thickness or 0
        density = row.density or 0
        quantity = row.quantity or 0
        used_quantity = row.used_quantity or 0
        used_weight = row.used_weight or 0
        used_meter = row.used_meter or 0
        rate_per_mtr = row.rate_per_mtr or 0

        if row.length and row.outer_diameter and row.thickness and row.density:
            od = row.outer_diameter
            inner_diameter = od - (2 * row.thickness)

            if inner_diameter > 0:
                row.weight_per_item = (pi * (((od / 2) ** 2) - ((inner_diameter / 2) ** 2)) * row.length * row.density) / 1000000
            else:
                row.weight_per_item = 0
        else:
            row.weight_per_item = 0

        row.balanced_quantity = ((row.quantity or 0) - (row.used_quantity or 0))
        row.actual_weight = (row.quantity or 1) * (row.weight_per_item or 0)
        row.available_weight = (row.actual_weight or 0) - (row.used_weight or 0)

        meter_per_item = length / 1000
        row.actual_meter = (quantity * meter_per_item)
        row.available_meter = max(row.actual_meter - used_meter,0)

        row.used_weight_percentage = (((row.actual_meter or 0) - (row.available_meter or 0)) / row.actual_meter * 100) if row.actual_meter else 0
        row.actual_weight_amount = (row.actual_weight or 0) * (row.rate_per_mtr or 0)
        row.available_weight_amount = (row.available_weight or 0) * (row.rate_per_mtr or 0)


    def calculate_rod(self, row):
        length = row.length or 0
        outer_diameter = row.outer_diameter or 0
        density = row.density or 0
        quantity = row.quantity or 0
        used_quantity = row.used_quantity or 0
        used_weight = row.used_weight or 0
        used_meter = row.used_meter or 0
        rate_per_mtr = row.rate_per_mtr or 0

        if row.length and row.outer_diameter and row.density:
            row.weight_per_item = (pi * ((row.outer_diameter / 2) ** 2) * row.length * row.density) / 1000000
        else:
            row.weight_per_item = 0

        row.balanced_quantity = ((row.quantity or 0) - (row.used_quantity or 0))
        row.actual_weight = (row.quantity or 1) * (row.weight_per_item or 0)
        row.available_weight = (row.actual_weight or 0) - (row.used_weight or 0)

        meter_per_item = length / 1000
        row.actual_meter = (quantity * meter_per_item)
        row.available_meter = max(row.actual_meter - used_meter,0)

        row.used_weight_percentage = (((row.actual_weight or 0) - (row.available_weight or 0)) / row.actual_weight * 100) if row.actual_weight else 0
        row.actual_weight_amount = (row.actual_weight or 0) * (row.rate_per_mtr or 0)
        row.available_weight_amount = (row.available_weight or 0) * (row.rate_per_mtr or 0)


@frappe.whitelist()
def calculate_stock_weights(doc):
    doc = frappe.parse_json(doc)
    stock_item = frappe.get_doc(doc)
    stock_item.calculate_stock_weights()
    return stock_item.as_dict()


def format_header(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

def auto_width(ws):
    for column in ws.columns:
        length = 0
        letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    length = max(length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = length + 5


def create_template_sheet(workbook, sheet_name, headers):
    ws = workbook.create_sheet(title=sheet_name)
    ws.append(headers)
    format_header(ws)
    auto_width(ws)
    return ws

def download_workbook(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "download"


@frappe.whitelist()
def download_template():
    template_type = frappe.form_dict.get("template_type")
    templates = {

        "plate": {
            "sheet": "Plates",
            "filename": "Stock_Plate_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Length", "Width", "Thickness", "Density",
                "Quantity", "Rate ₹ (Per Kg)", "Used Weight",
                "Weight Per Item", "Actual Weight", "Available Weight",
                "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "pipe": {
            "sheet": "Pipes",
            "filename": "Stock_Pipe_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Length", "Thickness", "Outer Diameter", "Density",
                "Quantity", "Used Quantity", "Rate ₹ (Per Mtr)",
                "Used Weight", "Used Meter" "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Meter", "Available Meter", "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "tube": {
            "sheet": "Tubes",
            "filename": "Stock_Tube_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Length", "Thickness", "Outer Diameter", "Density",
                "Quantity", "Used Quantity", "Rate ₹ (Per Mtr)",
                "Used Weight", "Used Meter", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Meter", "Available Meter", "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "rod": {
            "sheet": "Rods",
            "filename": "Stock_Rod_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Length", "Outer Diameter", "Density",
                "Quantity", "Used Quantity", "Rate ₹ (Per Mtr)",
                "Used Weight", "Used Meter", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Meter", "Available Meter", "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "flange": {
            "sheet": "Flanges",
            "filename": "Stock_Flange_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "NPS", "SCH", "Class",
                "Quantity", "Used Quantity", "Rate ₹ (Per Kg)",
                "Used Weight", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "welding": {
            "sheet": "Weldings",
            "filename": "Stock_Welding_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Size (mm)", "Heat No", "Batch No", "Make / Brand",
                "Quantity", "Used Quantity", "Rate ₹ (Per Kg)",
                "Used Weight", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "disc": {
            "sheet": "Disc",
            "filename": "Stock_Disc_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Size (mm)", "Heat No", "Batch No", "Make / Brand",
                "Quantity", "Used Quantity", "Rate ₹ (Per Kg)",
                "Used Weight", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "spares": {
            "sheet": "Spares",
            "filename": "Stock_Spares_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Size (mm)", "Heat No", "Batch No", "Make / Brand",
                "Quantity", "Used Quantity", "Rate ₹ (Per Kg)",
                "Used Weight", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        },

        "machinery": {
            "sheet": "Machinery",
            "filename": "Stock_Machinery_Template.xlsx",
            "headers": ["Category", "Type", "Material of Construction (MoC)",
                "Vendor", "Description", "Purchase Order No", "Status",
                "Size (mm)", "Heat No", "Batch No", "Make / Brand",
                "Quantity", "Used Quantity", "Rate ₹ (Per Kg)",
                "Used Weight", "Weight Per Item", "Balanced Quantity", "Actual Weight", "Available Weight",
                "Actual Weight Amount (₹)", "Available Weight Amount (₹)", "Remarks"]
        }
    }

    # -----------------------------
    # Overall Workbook
    # -----------------------------
    if template_type == "overall":
        workbook = Workbook()
        workbook.remove(workbook.active)

        for item_type in ["plate", "pipe", "tube", "rod", "flange", "welding", "disc", "spares", "machinery"]:
            template = templates[item_type]

            create_template_sheet(workbook, template["sheet"], template["headers"])
        download_workbook(workbook, "Stock_Overall_Template.xlsx")
        return

    # -----------------------------
    # Individual Workbook
    # -----------------------------
    if template_type not in templates:
        frappe.throw("Invalid template type")

    template = templates[template_type]
    workbook = Workbook()
    workbook.remove(workbook.active)

    create_template_sheet(workbook, template["sheet"], template["headers"])
    download_workbook(workbook, template["filename"])


DOWNLOAD_CONFIG = {
    "plates": {
        "sheet": "Plates",
        "table": "plates",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("Length", "length"),
            ("Width", "width"),
            ("Thickness", "thickness"),
            ("Density", "density"),
            ("Quantity", "quantity"),
            ("Rate ₹ (Per Kg)", "rate_per_kg"),
            ("Used Weight", "used_weight"),
            ("Weight Per Item", "weight_per_item"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Used Weight Percentage", "used_weight_percentage"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "pipes": {
        "sheet": "Pipes",
        "table": "pipes",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("Length", "length"),
            ("Thickness", "thickness"),
            ("Outer Diameter", "outer_diameter"),
            ("Density", "density"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Mtr)", "rate_per_mtr"),
            ("Used Weight", "used_weight"),
            ("Used Meter", "used_meter"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Meter", "actual_meter"),
            ("Available Meter", "available_meter"),
            ("Used Weight Percentage", "used_weight_percentage"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "tubes": {
        "sheet": "Tubes",
        "table": "tubes",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("Length", "length"),
            ("Thickness", "thickness"),
            ("Outer Diameter", "outer_diameter"),
            ("Density", "density"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Mtr)", "rate_per_mtr"),
            ("Used Weight", "used_weight"),
            ("Used Meter", "used_meter"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Meter", "actual_meter"),
            ("Available Meter", "available_meter"),
            ("Used Weight Percentage", "used_weight_percentage"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "rods": {
        "sheet": "Rods",
        "table": "rods",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("Length", "length"),
            ("Outer Diameter", "outer_diameter"),
            ("Density", "density"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Mtr)", "rate_per_mtr"),
            ("Used Weight", "used_weight"),
            ("Used Meter", "used_meter"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Meter", "actual_meter"),
            ("Available Meter", "available_meter"),
            ("Used Weight Percentage", "used_weight_percentage"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "flanges": {
        "sheet": "Flanges",
        "table": "flanges",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("NPS", "nps"),
            ("SCH", "sch"),
            ("Class", "class"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Kg)", "rate_per_kg"),
            ("Used Weight", "used_weight"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "welding": {
        "sheet": "Weldings",
        "table": "welding",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("Size (mm)", "size"),
            ("Heat No", "heat_no"),
            ("Batch No", "batch_no"),
            ("Make / Brand", "make_brand"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Kg)", "rate_per_kg"),
            ("Used Weight", "used_weight"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "disc": {
        "sheet": "Disc",
        "table": "disc",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("NPS", "nps"),
            ("SCH", "sch"),
            ("Class", "class"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Kg)", "rate_per_kg"),
            ("Used Weight", "used_weight"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "spares": {
        "sheet": "Spares",
        "table": "spares",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("NPS", "nps"),
            ("SCH", "sch"),
            ("Class", "class"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Kg)", "rate_per_kg"),
            ("Used Weight", "used_weight"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    },

    "machinery": {
        "sheet": "Machinery",
        "table": "machinery",
        "fields": [
            ("Category", "category"),
            ("Type", "type"),
            ("Material of Construction (MoC)", "moc"),
            ("Vendor", "vendor"),
            ("Description", "description"),
            ("Purchase Order No", "purchase_order_no"),
            ("Status", "status"),
            ("NPS", "nps"),
            ("SCH", "sch"),
            ("Class", "class"),
            ("Quantity", "quantity"),
            ("Used Quantity", "used_quantity"),
            ("Rate ₹ (Per Kg)", "rate_per_kg"),
            ("Used Weight", "used_weight"),
            ("Weight Per Item", "weight_per_item"),
            ("Balanced Quantity", "balanced_quantity"),
            ("Actual Weight", "actual_weight"),
            ("Available Weight", "available_weight"),
            ("Actual Weight Amount (₹)", "actual_weight_amount"),
            ("Available Weight Amount (₹)", "available_weight_amount"),
            ("Remarks", "remarks")
        ]
    }
}


@frappe.whitelist()
def download_stock_data(stock_item, download_type):
    try:
        # Convert JSON string from JS into Python list
        if isinstance(download_type, str):
            try:
                download_types = json.loads(download_type)
            except json.JSONDecodeError:
                download_types = [download_type]
        else:
            download_types = download_type

        # Make sure it is a list
        if not isinstance(download_types, list):
            download_types = [download_types]

        # Remove empty values and duplicates
        download_types = list(dict.fromkeys(str(item).lower().strip()
            for item in download_types if item))

        # Overall = all stock types
        if "overall" in download_types:
            download_types = ["plates", "pipes", "tubes", "rods", "flanges", "welding", "disc", "spares", "machinery"]

        if not download_types:
            frappe.throw("Please select at least one Download Type.")

        doc = frappe.get_doc("Stock Item", stock_item)
        wb = Workbook()

        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        for download_type in download_types:
            config = DOWNLOAD_CONFIG.get(download_type)

            if not config:
                frappe.throw(f"Invalid Download Type: {download_type}")

            ws = wb.create_sheet(title=config["sheet"])
            export_data(ws, doc, download_type)

        # Create Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if len(download_types) == 1:
            filename = (f"{stock_item}_{download_types[0]}.xlsx")
        else:
            filename = (f"{stock_item}_Stock_Data.xlsx")

        frappe.local.response.filename = filename
        frappe.local.response.filecontent = output.getvalue()
        frappe.local.response.type = "download"

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Stock Data Download Error")
        frappe.throw(str(e))


def export_data(ws, doc, download_type):
    config = DOWNLOAD_CONFIG.get(download_type)

    if not config:
        frappe.throw(f"Invalid Download Type: {download_type}")

    fields = config["fields"]

    # Header row
    ws.append([label for label, fieldname in fields])

    # Apply same header style as template
    format_header(ws)

    # Data rows
    for row in getattr(doc, config["table"], []):
        ws.append([getattr(row, fieldname, "")
            for label, fieldname in fields])

    # Auto adjust column width
    auto_width(ws)


@frappe.whitelist()
def upload_stock_excel(file_url, upload_type, stock_item):
    try:
        doc = frappe.get_doc("Stock Item", stock_item)
        upload_type = (upload_type or "").strip().lower()

        valid_types = ["plates", "pipes", "tubes", "rods", "flanges", "welding", "disc", "spares", "machinery", "overall"]

        if upload_type not in valid_types:
            frappe.throw(f"Invalid Upload Type: {upload_type}")

        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_path = file_doc.get_full_path()

        if upload_type == "overall":
            upload_overall_excel(file_path, doc)
        else:
            upload_single_excel(file_path, doc, upload_type)

        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "success": True,
            "message": "Excel uploaded successfully."
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Stock Excel Upload Error")
        frappe.throw(str(e))

def upload_overall_excel(file_path, doc):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_map = {
        "Plates": "plates",
        "Pipes": "pipes",
        "Tubes": "tubes",
        "Rods": "rods",
        "Flanges": "flanges",
        "Weldings": "welding",
        "Welding": "welding",
        "Disc": "disc",
        "Spares": "spares",
        "Machinery": "machinery"
    }

    found_sheet = False
    for sheet_name in workbook.sheetnames:

        if sheet_name not in sheet_map:
            continue

        upload_type = sheet_map[sheet_name]
        found_sheet = True
        upload_single_excel(file_path, doc, upload_type, sheet_name=sheet_name)

    if not found_sheet:
        frappe.throw("No valid stock sheets found in Overall Excel file.")
        

def upload_single_excel(file_path, doc, upload_type, sheet_name=None):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_map = {
        "plates": "Plates",
        "pipes": "Pipes",
        "tubes": "Tubes",
        "rods": "Rods",
        "flanges": "Flanges",
        "welding": "Weldings",
        "disc": "Disc",
        "spares": "Spares",
        "machinery": "Machinery"
    }

    if not sheet_name:
        sheet_name = sheet_map.get(upload_type)

    if not sheet_name:
        frappe.throw(f"Invalid Upload Type: {upload_type}")

    if sheet_name not in workbook.sheetnames:
        frappe.throw(f"'{sheet_name}' sheet not found in Excel file.")

    ws = workbook[sheet_name]
    upload_sheet_data(ws, doc, upload_type)


def upload_sheet_data(ws, doc, upload_type):
    table_map = {
        "plates": "plates",
        "pipes": "pipes",
        "tubes": "tubes",
        "rods": "rods",
        "flanges": "flanges",
        "welding": "welding",
        "disc": "disc",
        "spares": "spares",
        "machinery": "machinery"
    }

    table_field = table_map.get(upload_type)

    if not table_field:
        frappe.throw(f"Invalid Upload Type: {upload_type}")

    # -----------------------------------------
    # Read Excel headers
    # -----------------------------------------
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            headers.append("")
        else:
            headers.append(str(cell.value).strip())

    # -----------------------------------------
    # Excel header -> child field
    # -----------------------------------------
    field_map = {
        "Category": "category",
        "Type": "type",
        "Material of Construction (MoC)": "moc",
        "Vendor": "vendor",
        "Description": "description",
        "Purchase Order No": "purchase_order_no",
        "Purchase Order No": "purchase_order_no",
        "Status": "status",

        "Length": "length",
        "Width": "width",
        "Thickness": "thickness",
        "Outer Diameter": "outer_diameter",
        "Density": "density",

        "NPS": "nps",
        "SCH": "sch",
        "Class": "class",

        "Size (mm)": "size",
        "Heat No": "heat_no",
        "Batch No": "batch_no",
        "Make / Brand": "make_brand",

        "Quantity": "quantity",
        "Used Quantity": "used_quantity",

        "Rate ₹ (Per Kg)": "rate_per_kg",

        "Used Weight": "used_weight",
        "Used Meter": "used_meter",
        "Weight Per Item": "weight_per_item",
        "Balanced Quantity": "balanced_quantity",

        "Actual Weight": "actual_weight",
        "Available Weight": "available_weight",
        "Actual Meter": "actual_meter",
        "Available Meter": "available_meter",
        "Used Weight Percentage": "used_weight_percentage",

        "Actual Weight Amount (₹)": "actual_weight_amount",
        "Available Weight Amount (₹)": "available_weight_amount",

        "Remarks": "remarks"
    }

    # -----------------------------------------
    # Process Excel rows
    # -----------------------------------------
    for row in ws.iter_rows(min_row=2, values_only=True):

        # Skip completely empty rows
        if not any(value is not None for value in row):
            continue

        child = doc.append(table_field, {})

        # -------------------------------------
        # Fill child table fields
        # -------------------------------------
        for index, header in enumerate(headers):

            if not header:
                continue

            if index >= len(row):
                continue

            value = row[index]

            if value is None:
                continue

            frappe_field = field_map.get(header)

            if not frappe_field:
                continue

            # Only set field if it exists
            if frappe.get_meta(child.doctype).has_field(frappe_field):
                setattr(child, frappe_field, value)

        # -------------------------------------
        # Calculate values
        # -------------------------------------
        calculate_uploaded_row(child, upload_type)


def calculate_uploaded_row(child, upload_type):
    import math

    quantity = float(child.quantity or 0)
    used_quantity = float(getattr(child, "used_quantity", 0) or 0)
    rate = float(getattr(child, "rate_per_kg", 0) or 0)
    density = float(getattr(child, "density", 0) or 0)
    length = float(getattr(child, "length", 0) or 0)
    thickness = float(getattr(child, "thickness", 0) or 0)
    width = float(getattr(child, "width", 0) or 0)
    outer_diameter = float(getattr(child, "outer_diameter", 0) or 0)

    # -----------------------------------------
    # PLATES
    # -----------------------------------------
    if upload_type == "plates":
        weight_per_item = (length * width * thickness * density) / 1000000

    # -----------------------------------------
    # PIPES / TUBES
    # -----------------------------------------
    elif upload_type in ["pipes", "tubes"]:
        outer_radius = outer_diameter / 2
        inner_diameter = (outer_diameter - (2 * thickness))
        inner_radius = inner_diameter / 2
        weight_per_item = (math.pi * ((outer_radius ** 2) - (inner_radius ** 2)) * length * density) / 1000000

    # -----------------------------------------
    # RODS
    # -----------------------------------------
    elif upload_type == "rods":
        radius = outer_diameter / 2
        weight_per_item = (math.pi * (radius ** 2) * length * density) / 1000000

    else:
        return

    # -----------------------------------------
    # Weight Per Item
    # -----------------------------------------
    child.weight_per_item = weight_per_item

    # -----------------------------------------
    # Actual Weight
    # -----------------------------------------
    actual_weight = (weight_per_item * quantity)
    child.actual_weight = actual_weight

    # -----------------------------------------
    # Used Weight
    # -----------------------------------------
    if hasattr(child, "used_weight"):
        used_weight = (weight_per_item * used_quantity)
        child.used_weight = used_weight

    else:
        used_weight = 0

    # -----------------------------------------
    # Available Weight
    # -----------------------------------------
    available_weight = (actual_weight - used_weight)

    if hasattr(child, "available_weight"):
        child.available_weight = (available_weight)

    # -----------------------------------------
    # Balanced Quantity
    # -----------------------------------------
    if hasattr(child, "balanced_quantity"):
        child.balanced_quantity = (quantity - used_quantity)

    # -----------------------------------------
    # Actual Weight Amount
    # -----------------------------------------
    if hasattr(child, "actual_weight_amount"):
        child.actual_weight_amount = (actual_weight * rate)

    # -----------------------------------------
    # Available Weight Amount
    # -----------------------------------------
    if hasattr(child, "available_weight_amount"):
        child.available_weight_amount = (available_weight * rate)

def set_excel_value(child, row_data, excel_field, frappe_field):
    if excel_field in row_data:
        value = row_data[excel_field]

        if value is not None:
            setattr(child, frappe_field, value)