# Copyright (c) 2026, maze and contributors
# For license information, please see license.txt

# Copyright (c) 2026, maze and contributors
# For license information, please see license.txt

# import frappe
# from frappe.model.document import Document


# class WeeklyPOReport(Document):

#     @frappe.whitelist()
#     def send_weekly_po_mail(self):

#         frappe.sendmail(
#             recipients=["manager@company.com"],   # Change recipient
#             sender="msk312508@gmail.com",
#             subject=f"Weekly PO Report - {self.name}",
#             message="""
#                 Dear Team,<br><br>

#                 Please find the Weekly PO Report.<br><br>

#                 Regards,<br>
#                 DAPL ERP
#             """
#         )

#         return "Success"



import frappe
from frappe.model.document import Document
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO


class WeeklyPOReport(Document):

    @frappe.whitelist()
    def send_weekly_po_mail(self):

        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly PO Report"

        # Report Header
        ws["A1"] = "Weekly PO Report"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A3"] = "Week From"
        ws["B3"] = str(self.report_week_from_date or "")

        ws["A4"] = "Week To"
        ws["B4"] = str(self.report_week_to_date or "")

        ws["A5"] = "Prepared By"
        ws["B5"] = self.prepared_by or ""

        ws["A6"] = "Remarks"
        ws["B6"] = self.remarks or ""

        # Child Table Header
        row_no = 9

        headers = [
            "Order Number",
            "Order Type",
            "Supplier",
            "Job No",
            "Total",
            "Taxes",
            "Grand Total",
            "Description",
            "Status"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_no, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Child Table Data
        row_no += 1

        for d in self.items:
            ws.cell(row=row_no, column=1).value = d.order_number
            ws.cell(row=row_no, column=2).value = d.order_type
            ws.cell(row=row_no, column=3).value = d.supplier
            ws.cell(row=row_no, column=4).value = d.job_no
            ws.cell(row=row_no, column=5).value = d.total
            ws.cell(row=row_no, column=6).value = d.total_taxes_and_charges
            ws.cell(row=row_no, column=7).value = d.grand_total
            ws.cell(row=row_no, column=8).value = d.description
            ws.cell(row=row_no, column=9).value = d.status

            row_no += 1

        # Auto Width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 5, 50)

        # Create Excel in Memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        frappe.sendmail(
            recipients=["manager@company.com"],  # Change Recipient
            sender="msk312508@gmail.com",
            subject=f"Weekly PO Report - {self.name}",
            message=f"""
                Dear Team,<br><br>

                Please find attached the Weekly PO Report.<br><br>

                <b>Report:</b> {self.name}<br>
                <b>Week:</b> {self.report_week_from_date} to {self.report_week_to_date}<br><br>

                Regards,<br>
                DAPL ERP
            """,
            attachments=[
                {
                    "fname": f"{self.name}.xlsx",
                    "fcontent": output.getvalue()
                }
            ]
        )

        frappe.msgprint("Weekly PO Report mailed successfully.")