
# import frappe
# from frappe.utils import cint

# @frappe.whitelist()
# def get_dashboard_data(filters=None):

# 	# -----------------------------
# 	# SAFE FILTER PARSING
# 	# -----------------------------
# 	if not filters:
# 		filters = {}
# 	elif isinstance(filters, str):
# 		filters = frappe.parse_json(filters)

# 	# -----------------------------
# 	# CONDITION BUILDERS
# 	# -----------------------------
# 	# Conditions that only touch `Purchase Order` (po) / `Supplier` (sup).
# 	# with_item=True also applies item / item_group filters (requires poi join).
# 	# with_schedule_date=False skips the manual "Required By" exact-date filter,
# 	# used by the rolling next-7-days query below.
# 	def build_conditions(with_item=True, with_schedule_date=True):
# 		conditions = []
# 		if filters.get("supplier"):
# 			conditions.append("po.supplier = %(supplier)s")
# 		if filters.get("project"):
# 			conditions.append("po.project = %(project)s")
# 		if filters.get("order_type"):
# 			conditions.append("IFNULL(po.custom_order_type, 'Purchase Order') = %(order_type)s")
# 		if filters.get("status"):
# 			conditions.append("IFNULL(po.workflow_state, 'Draft') = %(status)s")
# 		if filters.get("transaction_date"):
# 			conditions.append("po.transaction_date = %(transaction_date)s")
# 		if with_schedule_date and filters.get("schedule_date"):
# 			conditions.append("po.schedule_date = %(schedule_date)s")
# 		if with_item:
# 			if filters.get("item_group"):
# 				conditions.append("poi.item_group = %(item_group)s")
# 			if filters.get("item"):
# 				conditions.append("poi.item_code = %(item)s")
# 		return conditions

# 	def build_where(conditions):
# 		return "WHERE " + " AND ".join(conditions) if conditions else ""

# 	where_po_only = build_where(build_conditions(with_item=False))
# 	where_with_item = build_where(build_conditions(with_item=True))

# 	# Rolling window: today through today + 7 days (inclusive), regardless of
# 	# the manual "Required By" filter - this is a fixed "what's due soon" view.
# 	where_upcoming = build_where(
# 		build_conditions(with_item=True, with_schedule_date=False)
# 		+ ["po.schedule_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 7 DAY)", "IFNULL(po.status, '') != 'Closed'"]
# 	)

# 	# -----------------------------
# 	# 1. ORDER TYPE COUNTS  (po only)
# 	# -----------------------------
# 	order_types = frappe.db.sql(f"""
# 		SELECT
# 			CASE
# 				WHEN po.custom_order_type IS NULL OR po.custom_order_type = '' THEN 'Others'
# 				ELSE po.custom_order_type
# 			END AS custom_order_type,
# 			COUNT(DISTINCT po.name) AS count,
# 			IFNULL(SUM(po.grand_total), 0) AS total_amount
# 		FROM `tabPurchase Order` po
# 		LEFT JOIN `tabSupplier` sup 
# 			ON sup.name = po.supplier

# 		{ where_po_only + (" AND " if where_po_only else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }
# 		GROUP BY 
# 			CASE
# 				WHEN po.custom_order_type IS NULL OR po.custom_order_type = '' THEN 'Others'
# 				ELSE po.custom_order_type
# 			END

# 		ORDER BY
# 			CASE 
# 				WHEN custom_order_type = 'Purchase Order' THEN 1
# 				WHEN custom_order_type = 'Work Order' THEN 2
# 				WHEN custom_order_type = 'Transport Order' THEN 3
# 				WHEN custom_order_type = 'Others' THEN 4
# 				ELSE 5
# 			END
# 	""", filters, as_dict=True)

# 	# -----------------------------
# 	# 2. TOP Tables
# 	# -----------------------------
# 	top_suppliers = frappe.db.sql(f"""
# 		SELECT po.supplier AS supplier,
# 			COUNT(DISTINCT po.name) AS count,
# 			IFNULL(SUM(poi.amount),0) AS total_amount,
# 			IFNULL(SUM(po.total_taxes_and_charges),0) AS total_taxes_and_charges

# 		FROM `tabPurchase Order` po

# 		INNER JOIN `tabPurchase Order Item` poi
# 			ON poi.parent = po.name

# 		LEFT JOIN `tabSupplier` sup
# 			ON sup.name = po.supplier

# 		{ where_with_item + (" AND " if where_with_item else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }

# 		GROUP BY po.supplier
# 		ORDER BY total_amount DESC
# 		LIMIT 10
# 	""", filters, as_dict=True)

# 	# Top Items (joins poi)
# 	top_items = frappe.db.sql(f"""
# 		SELECT poi.item_code AS item, poi.item_group,
# 			COUNT(DISTINCT po.name) AS order_count,
# 			IFNULL(SUM(poi.amount),0) AS basic_value,
# 			IFNULL(SUM(IFNULL(poi.igst_amount,0) + IFNULL(poi.cgst_amount,0) + IFNULL(poi.sgst_amount,0)),0) AS gst_value

# 		FROM `tabPurchase Order Item` poi
# 		INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
# 		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier

# 		{ where_with_item + (" AND " if where_with_item else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }
# 		GROUP BY poi.item_code, poi.item_group
# 		ORDER BY basic_value DESC
# 		LIMIT 10 """, filters, as_dict=True)

# 	# Top Projects
# 	top_projects = frappe.db.sql(f"""
# 		SELECT po.project, COUNT(DISTINCT po.name) AS count,

# 			/* Tag Name from Sales Order Item */
# 			IFNULL((
# 				SELECT GROUP_CONCAT(DISTINCT soi.item_code SEPARATOR ', ')
# 				FROM `tabSales Order Item` soi
# 				INNER JOIN `tabSales Order` so
# 					ON so.name = soi.parent
# 				WHERE soi.project = po.project
# 				AND so.docstatus = 1), '') AS tag_name,

# 			/* SO Basic Value */
# 			IFNULL((
# 				SELECT SUM(soi.amount)
# 				FROM `tabSales Order Item` soi
# 				INNER JOIN `tabSales Order` so
# 					ON so.name = soi.parent
# 				WHERE soi.project = po.project
# 				AND so.docstatus = 1), 0) AS so_basic_value,

# 			/* SO GST */
# 			IFNULL((
# 				SELECT SUM(IFNULL(soi.igst_amount,0) + IFNULL(soi.cgst_amount,0) + IFNULL(soi.sgst_amount,0))
# 				FROM `tabSales Order Item` soi
# 				INNER JOIN `tabSales Order` so
# 					ON so.name = soi.parent
# 				WHERE soi.project = po.project
# 				AND so.docstatus = 1), 0) AS so_gst_value,

# 			/* PO Basic Value */
# 			IFNULL(SUM(po.total),0) AS po_basic_value,

# 			/* PO Spend */
# 			IFNULL(SUM(po.total_taxes_and_charges),0) AS total_amount

# 		FROM `tabPurchase Order` po
# 		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier

# 		{ where_po_only + (" AND " if where_po_only else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }

# 		GROUP BY po.project
# 		ORDER BY total_amount DESC
# 		LIMIT 10 """, filters, as_dict=True)


# 	# -----------------------------
# 	# 3. WORKFLOW STATUS COUNTS (po only)
# 	# -----------------------------
# 	status_counts = frappe.db.sql(f"""
# 		SELECT IFNULL(po.workflow_state, 'Draft') AS workflow_state, COUNT(DISTINCT po.name) AS count
# 		FROM `tabPurchase Order` po
# 		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier
# 		{where_po_only}
# 		GROUP BY po.workflow_state
# 	""", filters, as_dict=True)

# 	# -----------------------------
# 	# 4. FULL PURCHASE ORDER LIST (respects all filters, incl. item/item_group/status)
# 	# -----------------------------
# 	limit = cint(filters.get("limit") or 20)
# 	offset = cint(filters.get("offset") or 0)

# 	full_po_list = frappe.db.sql(f"""
# 		SELECT po.name, po.supplier, po.project, po.transaction_date, po.schedule_date,
# 			IFNULL(po.custom_order_type, 'Purchase Order') AS custom_order_type,
# 			IFNULL(po.workflow_state, 'Draft') AS workflow_state,
# 			IFNULL(po.grand_total, 0) AS grand_total
# 		FROM `tabPurchase Order` po
# 		LEFT JOIN `tabPurchase Order Item` poi
# 			ON poi.parent = po.name
# 		LEFT JOIN `tabSupplier` sup
# 			ON sup.name = po.supplier
# 		{where_with_item}
# 		GROUP BY po.name
# 		ORDER BY po.transaction_date DESC, po.name DESC
# 		LIMIT %(limit)s OFFSET %(offset)s
# 	""", {
# 		**filters,
# 		"limit": limit,
# 		"offset": offset
# 	}, as_dict=True)

# 	# -----------------------------
# 	# 5. REQUIRED BY - NEXT 7 DAYS (rolling window from today)
# 	# -----------------------------
# 	upcoming_required_by = frappe.db.sql(f"""
# 		SELECT po.name, po.supplier, po.project, po.transaction_date, po.schedule_date,
# 			IFNULL(po.custom_order_type, 'Purchase Order') AS custom_order_type,
# 			IFNULL(po.workflow_state, 'Draft') AS workflow_state,
# 			IFNULL(po.grand_total, 0) AS grand_total
# 		FROM `tabPurchase Order` po
# 		LEFT JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
# 		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier
# 		{where_upcoming}
# 		GROUP BY po.name
# 		ORDER BY po.schedule_date ASC, po.name ASC
# 	""", filters, as_dict=True)

# 	# -----------------------------
# 	# 6. RFQ COUNT
# 	# -----------------------------
# 	rfq_conditions = []

# 	if filters.get("transaction_date"):
# 		rfq_conditions.append("transaction_date = %(transaction_date)s")

# 	where_rfq = "WHERE " + " AND ".join(rfq_conditions) if rfq_conditions else ""

# 	rfq_result = frappe.db.sql(f"""
# 		SELECT COUNT(name) AS count FROM `tabRequest for Quotation` {where_rfq} """, filters, as_dict=True)

# 	rfq_count = rfq_result[0].count if rfq_result else 0
	

# 	# -----------------------------
# 	# FINAL SAFE RESPONSE
# 	# -----------------------------
# 	return {
# 		"order_types": order_types or [],
# 		"top_suppliers": top_suppliers or [],
# 		"top_items": top_items or [],
# 		"top_projects": top_projects or [],
# 		# "top_item_groups": top_item_groups or [],
# 		"status_counts": status_counts or [],
# 		"full_po_list": full_po_list or [],
# 		"upcoming_required_by": upcoming_required_by or [],
# 		"rfq_count": rfq_count
# 	}




import frappe
from frappe.utils import cint

@frappe.whitelist()
def get_dashboard_data(filters=None):

	# -----------------------------
	# SAFE FILTER PARSING
	# -----------------------------
	if not filters:
		filters = {}
	elif isinstance(filters, str):
		filters = frappe.parse_json(filters)

	# -----------------------------
	# CONDITION BUILDERS
	# -----------------------------
	def build_conditions(with_item=True, with_schedule_date=True):
		conditions = []
		if filters.get("id"):
			filters["id_like"] = f"%{filters.get('id')}%"
			conditions.append("LOWER(po.name) LIKE LOWER(%(id_like)s)")
		if filters.get("supplier"):
			conditions.append("po.supplier = %(supplier)s")
		if filters.get("project"):
			conditions.append("po.project = %(project)s")
		if filters.get("order_type"):
			conditions.append("IFNULL(po.custom_order_type, 'Purchase Order') = %(order_type)s")
		if filters.get("status"):
			conditions.append("IFNULL(po.workflow_state, 'Draft') = %(status)s")
		if filters.get("transaction_date"):
			conditions.append("po.transaction_date = %(transaction_date)s")
		if with_schedule_date and filters.get("schedule_date"):
			conditions.append("po.schedule_date = %(schedule_date)s")
		if with_item:
			if filters.get("item_group"):
				conditions.append("poi.item_group = %(item_group)s")
			if filters.get("item"):
				conditions.append("poi.item_code = %(item)s")
		return conditions

	def build_where(conditions):
		return "WHERE " + " AND ".join(conditions) if conditions else ""

	where_po_only = build_where(build_conditions(with_item=False))
	where_with_item = build_where(build_conditions(with_item=True))

	where_upcoming = build_where(
		build_conditions(with_item=True, with_schedule_date=False)
		+ ["po.schedule_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 7 DAY)", "IFNULL(po.status, '') != 'Closed'"]
	)

	# -----------------------------
	# 1. ORDER TYPE COUNTS  (po only)
	# -----------------------------
	order_types = frappe.db.sql(f"""
		SELECT
			CASE
				WHEN po.custom_order_type IS NULL OR po.custom_order_type = '' THEN 'Others'
				ELSE po.custom_order_type
			END AS custom_order_type,
			COUNT(DISTINCT po.name) AS count,
			IFNULL(SUM(po.base_grand_total), 0) AS total_amount   -- was po.grand_total
		FROM `tabPurchase Order` po
		LEFT JOIN `tabSupplier` sup 
			ON sup.name = po.supplier

		{ where_po_only + (" AND " if where_po_only else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }
		GROUP BY 
			CASE
				WHEN po.custom_order_type IS NULL OR po.custom_order_type = '' THEN 'Others'
				ELSE po.custom_order_type
			END

		ORDER BY
			CASE 
				WHEN custom_order_type = 'Purchase Order' THEN 1
				WHEN custom_order_type = 'Work Order' THEN 2
				WHEN custom_order_type = 'Transport Order' THEN 3
				WHEN custom_order_type = 'Others' THEN 4
				ELSE 5
			END
	""", filters, as_dict=True)

	# -----------------------------
	# 2. TOP Tables
	# -----------------------------
	# top_suppliers = frappe.db.sql(f"""
	# 	SELECT po.supplier AS supplier,
	# 		COUNT(DISTINCT po.name) AS count,
	# 		IFNULL(SUM(po.total),0) AS total_amount,
	# 		IFNULL(SUM(po.base_total_taxes_and_charges),0) AS total_taxes_and_charges
	# 	FROM `tabPurchase Order` po

	# 	INNER JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
	# 	LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier
	# 	{ where_with_item + (" AND " if where_with_item else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }

	# 	GROUP BY po.supplier
	# 	ORDER BY total_amount DESC
	# 	LIMIT 10 """, filters, as_dict=True)

	top_suppliers = frappe.db.sql(f"""
		SELECT t.supplier AS supplier,
			COUNT(DISTINCT t.name) AS count,
			IFNULL(SUM(t.total), 0) AS total_amount,
			IFNULL(SUM(t.base_total_taxes_and_charges), 0) AS total_taxes_and_charges
		FROM (
			SELECT DISTINCT
				po.name AS name,
				po.supplier AS supplier,
				po.total AS total,
				po.base_total_taxes_and_charges AS base_total_taxes_and_charges
			FROM `tabPurchase Order` po
			INNER JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
			LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier
			{ where_with_item + (" AND " if where_with_item else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }
		) t
		GROUP BY t.supplier
		ORDER BY total_amount DESC
		LIMIT 10
	""", filters, as_dict=True)

	# Top Items (joins poi)
	top_items = frappe.db.sql(f"""
		SELECT poi.item_code AS item, poi.item_group,
			COUNT(DISTINCT po.name) AS order_count,
			IFNULL(SUM(poi.base_amount),0) AS basic_value,
			IFNULL(SUM(
				(IFNULL(poi.igst_amount,0) + IFNULL(poi.cgst_amount,0) + IFNULL(poi.sgst_amount,0))
				* IFNULL(po.conversion_rate, 1)),0) AS gst_value

		FROM `tabPurchase Order Item` poi
		INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier

		{ where_with_item + (" AND " if where_with_item else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }
		GROUP BY poi.item_code, poi.item_group
		ORDER BY basic_value DESC
		LIMIT 10 """, filters, as_dict=True)

	# Top Projects
	top_projects = frappe.db.sql(f"""
		SELECT po.project, COUNT(DISTINCT po.name) AS count,

			/* Tag Name from Sales Order Item */
			IFNULL((
				SELECT GROUP_CONCAT(DISTINCT soi.item_code SEPARATOR ', ')
				FROM `tabSales Order Item` soi
				INNER JOIN `tabSales Order` so
					ON so.name = soi.parent
				WHERE soi.project = po.project
				AND so.docstatus = 1), '') AS tag_name,

			/* SO Basic Value */
			IFNULL((
				SELECT SUM(soi.base_amount)                 -- was soi.amount
				FROM `tabSales Order Item` soi
				INNER JOIN `tabSales Order` so
					ON so.name = soi.parent
				WHERE soi.project = po.project
				AND so.docstatus = 1), 0) AS so_basic_value,

			/* SO GST */
			IFNULL((
				SELECT SUM(
					(IFNULL(soi.igst_amount,0) + IFNULL(soi.cgst_amount,0) + IFNULL(soi.sgst_amount,0))
					* IFNULL(so.conversion_rate, 1)
				)
				FROM `tabSales Order Item` soi
				INNER JOIN `tabSales Order` so
					ON so.name = soi.parent
				WHERE soi.project = po.project
				AND so.docstatus = 1), 0) AS so_gst_value,

			/* PO Basic Value */
			IFNULL(SUM(po.base_total),0) AS po_basic_value,   -- was po.total

			/* PO Spend */
			IFNULL(SUM(po.base_total_taxes_and_charges),0) AS total_amount   -- was po.total_taxes_and_charges

		FROM `tabPurchase Order` po
		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier

		{ where_po_only + (" AND " if where_po_only else "WHERE ") + "IFNULL(po.workflow_state,'Draft') != 'Cancelled'" }

		GROUP BY po.project
		ORDER BY total_amount DESC
		LIMIT 10 """, filters, as_dict=True)


	# -----------------------------
	# 3. WORKFLOW STATUS COUNTS (po only) — no money fields, unchanged
	# -----------------------------
	status_counts = frappe.db.sql(f"""
		SELECT IFNULL(po.workflow_state, 'Draft') AS workflow_state, COUNT(DISTINCT po.name) AS count
		FROM `tabPurchase Order` po
		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier
		{where_po_only}
		GROUP BY po.workflow_state
	""", filters, as_dict=True)

	# -----------------------------
	# 4. FULL PURCHASE ORDER LIST
	# -----------------------------
	limit = cint(filters.get("limit") or 20)
	offset = cint(filters.get("offset") or 0)

	full_po_list = frappe.db.sql(f"""
		SELECT po.name, po.supplier, po.project, po.transaction_date, po.schedule_date,
			IFNULL(po.custom_order_type, 'Purchase Order') AS custom_order_type,
			IFNULL(po.workflow_state, 'Draft') AS workflow_state,
			IFNULL(po.base_grand_total, 0) AS grand_total    -- was po.grand_total
		FROM `tabPurchase Order` po
		LEFT JOIN `tabPurchase Order Item` poi
			ON poi.parent = po.name
		LEFT JOIN `tabSupplier` sup
			ON sup.name = po.supplier
		{where_with_item}
		GROUP BY po.name
		ORDER BY po.transaction_date DESC, po.name DESC
		LIMIT %(limit)s OFFSET %(offset)s
	""", {
		**filters,
		"limit": limit,
		"offset": offset
	}, as_dict=True)

	# -----------------------------
	# 5. REQUIRED BY - NEXT 7 DAYS
	# -----------------------------
	upcoming_required_by = frappe.db.sql(f"""
		SELECT po.name, po.supplier, po.project, po.transaction_date, po.schedule_date,
			IFNULL(po.custom_order_type, 'Purchase Order') AS custom_order_type,
			IFNULL(po.workflow_state, 'Draft') AS workflow_state,
			IFNULL(po.base_grand_total, 0) AS grand_total    -- was po.grand_total
		FROM `tabPurchase Order` po
		LEFT JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
		LEFT JOIN `tabSupplier` sup ON sup.name = po.supplier
		{where_upcoming}
		GROUP BY po.name
		ORDER BY po.schedule_date ASC, po.name ASC
	""", filters, as_dict=True)

	# -----------------------------
	# 6. RFQ COUNT — no money fields, unchanged
	# -----------------------------
	rfq_conditions = []

	if filters.get("transaction_date"):
		rfq_conditions.append("transaction_date = %(transaction_date)s")

	where_rfq = "WHERE " + " AND ".join(rfq_conditions) if rfq_conditions else ""

	rfq_result = frappe.db.sql(f"""
		SELECT COUNT(name) AS count FROM `tabRequest for Quotation` {where_rfq} """, filters, as_dict=True)

	rfq_count = rfq_result[0].count if rfq_result else 0
	

	# -----------------------------
	# FINAL SAFE RESPONSE
	# -----------------------------
	return {
		"order_types": order_types or [],
		"top_suppliers": top_suppliers or [],
		"top_items": top_items or [],
		"top_projects": top_projects or [],
		"status_counts": status_counts or [],
		"full_po_list": full_po_list or [],
		"upcoming_required_by": upcoming_required_by or [],
		"rfq_count": rfq_count
	}


@frappe.whitelist()
def download_purchase_excel(filters=None):

	import io
	import openpyxl

	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	from openpyxl.utils import get_column_letter

	# -------------------------------------------------
	# SAFE FILTER PARSING
	# -------------------------------------------------

	if not filters:
		filters = {}

	elif isinstance(filters, str):
		filters = frappe.parse_json(filters)

	# -------------------------------------------------
	# BUILD CONDITIONS
	# -------------------------------------------------

	conditions = []
	if filters.get("id"):
		filters["id_like"] = f"%{filters.get('id')}%"
		conditions.append("LOWER(po.name) LIKE LOWER(%(id_like)s)")
			
	if filters.get("supplier"):
		conditions.append("po.supplier = %(supplier)s")

	if filters.get("project"):
		conditions.append("po.project = %(project)s")

	if filters.get("order_type"):
		conditions.append("IFNULL(po.custom_order_type, 'Purchase Order') = %(order_type)s")

	if filters.get("status"):
		conditions.append("IFNULL(po.workflow_state, 'Draft') = %(status)s")

	if filters.get("transaction_date"):
		conditions.append("po.transaction_date = %(transaction_date)s")

	if filters.get("schedule_date"):
		conditions.append("po.schedule_date = %(schedule_date)s")

	if filters.get("item"):
		conditions.append("poi.item_code = %(item)s")

	if filters.get("item_group"):
		conditions.append("poi.item_group = %(item_group)s")

	where_clause = ""

	if conditions:
		where_clause = "WHERE " + " AND ".join(conditions)

	# -------------------------------------------------
	# GET PURCHASE ORDER + ITEM DATA
	# -------------------------------------------------

	data = frappe.db.sql(
		f"""
		SELECT
			po.name AS purchase_order_no,
			po.supplier,
			po.project,
			IFNULL(po.custom_order_type, 'Purchase Order') AS custom_order_type,

			poi.item_code,
			poi.item_group,
			poi.description,

			poi.custom_material_type,
			poi.custom_length,
			poi.custom_width,
			poi.custom_thickness,
			poi.custom_outer_diameter,
			poi.custom_inner_diameter,
			poi.custom_density,

			poi.qty,
			poi.uom,
			poi.rate, poi.custom_rate_per_kg,
			poi.amount

		FROM `tabPurchase Order` po

		INNER JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
		{where_clause}

		ORDER BY po.transaction_date DESC, po.name DESC, poi.idx ASC""", filters, as_dict=True)

	# -------------------------------------------------
	# NO DATA
	# -------------------------------------------------

	if not data:
		frappe.throw("No Purchase Order data found for the selected filters.")

	# -------------------------------------------------
	# CREATE WORKBOOK
	# -------------------------------------------------

	wb = Workbook()
	ws = wb.active
	ws.title = "Purchase Orders"

	# -------------------------------------------------
	# HEADER
	# -------------------------------------------------

	headers = ["Purchase Order No", "Supplier", "Project", "Order Type",
		"Item Code", "Item Group", "Description", "Material Type",
		"Length", "Width", "Thickness", "Outer Diameter", "Inner Diameter", "Density",
		"Qty", "UOM", "Rate", "Rate (Per Kg / Per Mtr)", "Amount"]

	ws.append(headers)

	# -------------------------------------------------
	# HEADER STYLE
	# -------------------------------------------------

	header_fill = PatternFill(fill_type="solid", fgColor="22C55E")
	header_font = Font(bold=True, color="FFFFFF", size=11)
	header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	thin_border = Border(
		left=Side(style="thin", color="D1D5DB"),
		right=Side(style="thin", color="D1D5DB"),
		top=Side(style="thin", color="D1D5DB"),
		bottom=Side(style="thin", color="D1D5DB")
	)

	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = header_alignment
		cell.border = thin_border

	# Header height
	ws.row_dimensions[1].height = 30
	ws.freeze_panes = "A2"

	# -------------------------------------------------
	# WRITE DATA
	# -------------------------------------------------

	previous_po = None
	for row in data:
		current_po = row.get("purchase_order_no")

		# Repeat Purchase Order parent details for EVERY item row
		po_no = row.get("purchase_order_no") or ""
		supplier = row.get("supplier") or ""
		project = row.get("project") or ""
		order_type = row.get("custom_order_type") or "Purchase Order"

		excel_row = [
			po_no,
			supplier,
			project,
			order_type,

			row.get("item_code") or "",
			row.get("item_group") or "",
			# Description - remove HTML tags
			# frappe.utils.strip_html(row.get("description") or "").strip(),
			frappe.utils.strip_html((row.get("description") or "").replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")).strip(),

			row.get("custom_material_type") or "",
			row.get("custom_length"),
			row.get("custom_width"),
			row.get("custom_thickness"),
			row.get("custom_outer_diameter"),
			row.get("custom_inner_diameter"),
			row.get("custom_density"),

			row.get("qty"),
			row.get("uom") or "",
			row.get("rate"),
			row.get("custom_rate_per_kg"),
			row.get("amount")
		]

		ws.append(excel_row)
		previous_po = current_po

	# -------------------------------------------------
	# FORMAT DATA CELLS
	# -------------------------------------------------

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
		for cell in row:
			cell.border = thin_border
			cell.alignment = Alignment(vertical="center")

	# -------------------------------------------------
	# NUMBER FORMATTING
	# -------------------------------------------------

	# Length
	for row in range(2, ws.max_row + 1):
		for col in [
			9,   # Length
			10,  # Width
			11,  # Thickness
			12,  # Outer Diameter
			13,  # Inner Diameter
			14   # Density
		]:

			ws.cell(row=row, column=col).number_format = "0.000"

		# Qty
		ws.cell(row=row, column=15).number_format = "0.000"

		# Rate
		ws.cell(row=row, column=17).number_format = '#,##0.00'

		# Rate (Per Kg / Per Mtr)
		ws.cell(row=row, column=18).number_format = '#,##0.00'

		# Amount
		ws.cell(row=row, column=19).number_format = '#,##0.00'

	# -------------------------------------------------
	# COLUMN WIDTHS
	# -------------------------------------------------

	column_widths = {
		"A": 20,
		"B": 28,
		"C": 25,
		"D": 20,

		"E": 20,
		"F": 25,
		"G": 45,

		"H": 22,

		"I": 14,
		"J": 14,
		"K": 14,
		"L": 18,
		"M": 18,
		"N": 14,

		"O": 12,
		"P": 12,
		"Q": 15,
		"R": 18,
		"S": 24
	}

	for column, width in column_widths.items():
		ws.column_dimensions[column].width = width

	# -------------------------------------------------
	# ALTERNATE ROW STYLE
	# -------------------------------------------------

	alternate_fill = PatternFill(fill_type="solid", fgColor="F0FDF4")
	for row_number in range(2, ws.max_row + 1):
		if row_number % 2 == 0:
			for cell in ws[row_number]:
				cell.fill = alternate_fill

	ws.auto_filter.ref = ws.dimensions
	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	# -------------------------------------------------
	# CREATE FILE IN FRAPPE
	# -------------------------------------------------

	filename = ("Purchase_Order_Details_" + frappe.utils.now_datetime().strftime("%Y%m%d_%H%M%S") + ".xlsx")

	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": filename,
		"content": output.getvalue(),
		"is_private": 1
	})

	file_doc.save(ignore_permissions=True)
	return file_doc.file_url